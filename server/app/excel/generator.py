"""
Excel report generator.

Creates a professionally formatted .xlsx workbook with three sheets:
1. Account Details
2. Transaction Ledger
3. Analytics
"""

import logging
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
_KEY_FONT = Font(name="Calibri", bold=True, size=11)
_VALUE_FONT = Font(name="Calibri", size=11)
_ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_CURRENCY_FMT = '#,##0.00'


class ExcelGenerator:
    """Generate a multi-sheet Excel report from parsed statement data."""

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

    def generate(
        self,
        account_details: Dict[str, Any],
        transactions: List[Dict[str, Any]],
        analytics: Dict[str, Any],
        output_path: str,
    ) -> str:
        """
        Create the Excel workbook and write it to ``output_path``.

        Args:
            account_details: Parsed account metadata dict.
            transactions: List of categorised transaction dicts.
            analytics: Output from ``AnalyticsEngine.generate_full_analytics``.
            output_path: Absolute file path for the .xlsx file.

        Returns:
            The ``output_path`` for convenience.
        """
        wb = Workbook()

        # Sheet 1 – Account Details
        ws1 = wb.active
        ws1.title = "Account Details"
        self._write_account_details(ws1, account_details)

        # Sheet 2 – Transaction Ledger
        ws2 = wb.create_sheet("Transaction Ledger")
        self._write_transactions(ws2, transactions)

        # Sheet 3 – Analytics
        ws3 = wb.create_sheet("Analytics")
        self._write_analytics(ws3, analytics)

        wb.save(output_path)
        logger.info("Excel report saved to %s", output_path)
        return output_path

    # ------------------------------------------------------------------ #
    # Sheet 1 – Account Details
    # ------------------------------------------------------------------ #

    def _write_account_details(self, ws, details: Dict[str, Any]) -> None:
        """Write key-value account details in two columns."""
        ws.sheet_properties.tabColor = "4472C4"

        fields = [
            ("Account Holder Name", details.get("account_holder_name", "")),
            ("Account Number", details.get("account_number", "")),
            ("Bank Name", details.get("bank_name", "HDFC Bank")),
            ("Branch", details.get("branch", "")),
            ("IFSC Code", details.get("ifsc", "")),
            ("Statement Period", details.get("statement_period", "")),
            ("Opening Balance", details.get("opening_balance", 0.0)),
            ("Closing Balance", details.get("closing_balance", 0.0)),
            ("Total Credits", details.get("total_credits", 0.0)),
            ("Total Debits", details.get("total_debits", 0.0)),
        ]

        # Title row
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = "HDFC Bank Statement – Account Details"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        title_cell.alignment = _CENTER

        row = 3
        for key, value in fields:
            key_cell = ws.cell(row=row, column=1, value=key)
            key_cell.font = _KEY_FONT
            key_cell.border = _THIN_BORDER

            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.font = _VALUE_FONT
            val_cell.border = _THIN_BORDER
            if isinstance(value, (int, float)):
                val_cell.number_format = _CURRENCY_FMT

            row += 1

        # Auto-width
        self._auto_width(ws, columns=[1, 2])

    # ------------------------------------------------------------------ #
    # Sheet 2 – Transaction Ledger
    # ------------------------------------------------------------------ #

    def _write_transactions(self, ws, transactions: List[Dict[str, Any]]) -> None:
        """Write the full transaction table with formatting."""
        ws.sheet_properties.tabColor = "70AD47"

        headers = [
            "Date",
            "Value Date",
            "Description",
            "Cheque/Ref No",
            "Deposits (CR)",
            "Withdrawals (DR)",
            "Running Balance",
            "Category",
        ]

        # Header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

        # Data rows
        for row_idx, txn in enumerate(transactions, start=2):
            values = [
                txn.get("date", ""),
                txn.get("value_date", ""),
                txn.get("description", ""),
                txn.get("cheque_ref_no", ""),
                txn.get("credit", 0.0),
                txn.get("debit", 0.0),
                txn.get("running_balance", 0.0),
                txn.get("category", "Other"),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = _THIN_BORDER
                cell.font = _VALUE_FONT

                # Currency formatting for amount columns (5, 6, 7)
                if col_idx in (5, 6, 7) and isinstance(value, (int, float)):
                    cell.number_format = _CURRENCY_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx == 3:
                    cell.alignment = _LEFT
                else:
                    cell.alignment = _CENTER

            # Alternate row shading
            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = _ALT_ROW_FILL

        # Auto-width
        self._auto_width(ws, columns=list(range(1, len(headers) + 1)))

        # Freeze top row
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # Sheet 3 – Analytics
    # ------------------------------------------------------------------ #

    def _write_analytics(self, ws, analytics: Dict[str, Any]) -> None:
        """Write all analytics sections with professional formatting."""
        ws.sheet_properties.tabColor = "ED7D31"
        current_row = 1

        # Section 1 – Category Summary
        current_row = self._write_section_header(ws, current_row, "Category Summary")
        cat_summary = analytics.get("category_summary", [])
        if cat_summary:
            cat_headers = ["Category", "Transactions", "Total Debit", "Total Credit"]
            current_row = self._write_table(ws, current_row, cat_headers, [
                [c["category"], c["transaction_count"], c["total_debit"], c["total_credit"]]
                for c in cat_summary
            ], currency_cols=[3, 4])
        current_row += 2

        # Section 2 – Monthly Analysis
        current_row = self._write_section_header(ws, current_row, "Monthly Analysis")
        monthly = analytics.get("monthly_analysis", [])
        if monthly:
            monthly_headers = ["Month", "Inflows (CR)", "Outflows (DR)", "Net"]
            current_row = self._write_table(ws, current_row, monthly_headers, [
                [m["month"], m["inflows"], m["outflows"], m["net"]]
                for m in monthly
            ], currency_cols=[2, 3, 4])
        current_row += 2

        # Section 3 – Top 5 Largest Transactions
        current_row = self._write_section_header(ws, current_row, "Top 5 Largest Transactions")
        top = analytics.get("top_transactions", [])
        if top:
            top_headers = ["Date", "Description", "Credit", "Debit", "Amount", "Category"]
            current_row = self._write_table(ws, current_row, top_headers, [
                [t["date"], t["description"], t["credit"], t["debit"], t["amount"], t["category"]]
                for t in top
            ], currency_cols=[3, 4, 5])
        current_row += 2

        # Section 4 – Salary Detection
        current_row = self._write_section_header(ws, current_row, "Salary Detection")
        salary = analytics.get("salary_detection", [])
        if salary:
            sal_headers = ["Month", "Date", "Amount", "Description"]
            current_row = self._write_table(ws, current_row, sal_headers, [
                [s["month"], s["date"], s["amount"], s["description"]]
                for s in salary
            ], currency_cols=[3])
        else:
            ws.cell(row=current_row, column=1, value="No recurring salary pattern detected.").font = _VALUE_FONT
            current_row += 1
        current_row += 2

        # Section 5 – EMI Detection
        current_row = self._write_section_header(ws, current_row, "EMI Detection")
        emi = analytics.get("emi_detection", [])
        if emi:
            emi_headers = ["Month", "Date", "Amount", "Description"]
            current_row = self._write_table(ws, current_row, emi_headers, [
                [e["month"], e["date"], e["amount"], e["description"]]
                for e in emi
            ], currency_cols=[3])
        else:
            ws.cell(row=current_row, column=1, value="No recurring EMI pattern detected.").font = _VALUE_FONT
            current_row += 1
        current_row += 2

        # Section 6 – Categorization Statistics
        current_row = self._write_section_header(ws, current_row, "Categorization Statistics")
        stats = analytics.get("categorization_stats", {})
        stat_rows = [
            ("Total Transactions", stats.get("total", 0)),
            ("Categorized", stats.get("categorized_count", 0)),
            ("Uncategorized", stats.get("uncategorized_count", 0)),
            ("Categorized %", f"{stats.get('categorized_pct', 0):.2f}%"),
            ("Uncategorized %", f"{stats.get('uncategorized_pct', 0):.2f}%"),
        ]
        for key, val in stat_rows:
            ws.cell(row=current_row, column=1, value=key).font = _KEY_FONT
            ws.cell(row=current_row, column=2, value=val).font = _VALUE_FONT
            current_row += 1

        # Auto-width all used columns
        self._auto_width(ws, columns=list(range(1, 7)))

    # ------------------------------------------------------------------ #
    # Formatting helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _write_section_header(ws, row: int, title: str) -> int:
        """Write a bold section title and return the next row."""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = _SECTION_FONT
        return row + 1

    @staticmethod
    def _write_table(
        ws,
        start_row: int,
        headers: List[str],
        data: List[List[Any]],
        currency_cols: List[int] | None = None,
    ) -> int:
        """
        Write a table with header + data rows and return the next available row.

        ``currency_cols`` are 1-indexed column numbers that should be formatted
        as currency.
        """
        currency_cols = currency_cols or []

        # Header
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

        # Data
        for row_offset, row_data in enumerate(data, start=1):
            current_row = start_row + row_offset
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = _VALUE_FONT
                cell.border = _THIN_BORDER
                if col_idx in currency_cols and isinstance(value, (int, float)):
                    cell.number_format = _CURRENCY_FMT
                    cell.alignment = Alignment(horizontal="right")

            # Alternate shading
            if row_offset % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=current_row, column=col_idx).fill = _ALT_ROW_FILL

        return start_row + len(data) + 1

    @staticmethod
    def _auto_width(ws, columns: List[int], min_width: int = 10, max_width: int = 50) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx in columns:
            col_letter = get_column_letter(col_idx)
            max_len = min_width
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell_len = len(str(cell.value))
                        max_len = max(max_len, min(cell_len + 2, max_width))
            ws.column_dimensions[col_letter].width = max_len
